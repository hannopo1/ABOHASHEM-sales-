#!/usr/bin/env python3
"""Export every analysis output into one multi-sheet Excel workbook for
non-technical review (Board/Investor/Bank audiences who want Excel, not CSV/JSON)."""
import json
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
P = ROOT / "data" / "processed"
OUT = ROOT / "reports" / "ABOHASHEM_full_analysis.xlsx"

HEADER_FILL = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F2A44")


def style_sheet(ws, df, rtl=True, freeze="A2"):
    ws.sheet_view.rightToLeft = rtl
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def write_df(writer, df, sheet_name, rtl=True):
    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    ws = writer.sheets[sheet_name[:31]]
    style_sheet(ws, df, rtl=rtl)


def dict_to_df(d, key_label="المؤشر", val_label="القيمة"):
    rows = []
    for k, v in d.items():
        if isinstance(v, (dict, list)):
            continue
        rows.append({key_label: k, val_label: v})
    return pd.DataFrame(rows)


def main():
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        # 1. Executive summary
        eda = json.load(open(P / "eda_summary.json", encoding="utf-8"))
        fin = json.load(open(P / "financial_analysis.json", encoding="utf-8"))
        dq = json.load(open(P / "data_quality_metrics.json", encoding="utf-8"))
        summary_rows = [
            ("إجمالي الإيراد (18 شهرًا)", fin["total_revenue_egp"]),
            ("إيراد آخر 12 شهرًا (Trailing 12M)", fin["trailing_12m_revenue_egp"]),
            ("متوسط الإيراد الشهري (آخر 12 شهرًا)", fin["avg_monthly_revenue_t12_egp"]),
            ("القيمة الاسمية الإجمالية (الكمية × السعر)", fin["gross_list_value_egp"]),
            ("إجمالي الاستقطاعات التجارية", fin["aggregate_deduction_value_egp"]),
            ("نسبة الاستقطاع الإجمالية %", fin["aggregate_deduction_rate_pct"]),
            ("قيمة مبيعات البونص المقدَّرة", fin["bonus_estimated_value_at_asp_egp"]),
            ("عدد العملاء", eda["n_customers"]),
            ("عدد الأصناف المباعة فعليًا", eda["n_items"]),
            ("عدد العلامات التجارية", 3),
            ("حصة أعلى 10 عملاء من الإيراد %", fin["top10_customer_share_pct"]),
            ("حصة أعلى 10 أصناف من الإيراد %", fin["top10_item_share_pct"]),
            ("مؤشر تركّز العملاء HHI", fin["hhi_customers"]),
            ("مؤشر تركّز العلامات التجارية HHI", fin["hhi_brands"]),
            ("مؤشر تركّز الأصناف HHI", fin["hhi_items"]),
            ("صافي رصيد المديونية (لقطة 2026/7/4)", fin["ar_total_net_balance_egp"]),
            ("أيام الذمم المدينة التقريبية (DSO)", fin["dso_proxy_days"]),
            ("حصة أعلى 10 مدينين من إجمالي المديونية %", fin["ar_top10_debtor_share_pct"]),
            ("إجمالي الفواتير المُحلَّلة", dq["n_invoices"]),
            ("إجمالي بنود الفواتير", dq["n_rows"]),
            ("فواتير بعدم تطابق (>1 ج.م)", dq["n_invoices_reconciliation_mismatch_over_1egp"]),
            ("هامش الربح / EBITDA / صافي الربح", "غير متاح — لا توجد بيانات تكلفة في الملفات المرفوعة"),
        ]
        df_summary = pd.DataFrame(summary_rows, columns=["المؤشر", "القيمة"])
        write_df(writer, df_summary, "0_الملخص التنفيذي")

        # 2. Monthly series
        write_df(writer, pd.read_csv(P / "eda_monthly_series.csv"), "1_السلسلة الشهرية")

        # 3. Customer ABC/XYZ
        write_df(writer, pd.read_csv(P / "eda_customer_pareto_abc.csv"), "2_تحليل العملاء ABC-XYZ")

        # 4. Item ABC/XYZ
        write_df(writer, pd.read_csv(P / "eda_item_abc_xyz.csv"), "3_تحليل الأصناف ABC-XYZ")

        # 5. Brand summary + ASP by brand-month
        write_df(writer, pd.read_csv(P / "eda_brand_summary.csv"), "4_أداء العلامات التجارية")
        write_df(writer, pd.read_csv(P / "eda_asp_by_brand_month.csv"), "5_متوسط السعر شهريًا للعلامة")

        # 6. Hierarchy tree (customer -> brand -> item)
        write_df(writer, pd.read_csv(P / "eda_hierarchy_customer_brand_item.csv"), "6_الشجرة الهرمية عميل-علامة-صنف")

        # 7. Item & customer dimensions
        write_df(writer, pd.read_csv(P / "dim_items.csv"), "7_دليل الأصناف والبراند والكرتونة")
        write_df(writer, pd.read_csv(P / "dim_customers.csv"), "8_دليل العملاء")

        # 8. AR balances
        ar = pd.read_csv(P / "ar_customer_balances_2026-07-04.csv")
        ar["net_balance"] = ar["debit"] - ar["credit"]
        write_df(writer, ar, "9_أرصدة المديونية 2026-07-04")

        # 9. Forecast - company level
        fc = json.load(open(P / "forecast_results.json", encoding="utf-8"))
        df_fc_company = pd.DataFrame(fc["forecast_company_revenue"])
        write_df(writer, df_fc_company, "10_توقع الشركة 7 أشهر")

        df_cv = pd.DataFrame([
            {"النموذج": k, **v} for k, v in fc["cv_summary"].items() if v
        ]).sort_values("rmse")
        write_df(writer, df_cv, "11_مقارنة نماذج التنبؤ")

        # 10. Forecast - brand/customer/item disaggregated
        fcd = json.load(open(P / "forecast_disaggregated.json", encoding="utf-8"))
        brand_rows = []
        for b, v in fcd["brands"].items():
            for r in v["forecast"]:
                brand_rows.append({"العلامة التجارية": b, **r})
        write_df(writer, pd.DataFrame(brand_rows), "12_توقع العلامات التجارية")

        cust_rows = []
        for code, v in fcd["top_customers"].items():
            for r in v["forecast"]:
                cust_rows.append({"كود العميل": code, "اسم العميل": v["name"], **r})
        write_df(writer, pd.DataFrame(cust_rows), "13_توقع أعلى 10 عملاء")

        item_rows = []
        for name, v in fcd["top_items"].items():
            for r in v["forecast"]:
                item_rows.append({"الصنف": name, **r})
        write_df(writer, pd.DataFrame(item_rows), "14_توقع أعلى 10 أصناف")

        # 11. Time-series diagnostics (flattened, numeric parts only)
        ts = json.load(open(P / "timeseries_diagnostics.json", encoding="utf-8"))
        ts_rows = [
            ("ADF - المستوى - الإحصائية", ts["adf_level"]["stat"]),
            ("ADF - المستوى - القيمة الاحتمالية", ts["adf_level"]["pvalue"]),
            ("ADF - الفرق الأول - الإحصائية", ts["adf_first_diff"]["stat"]),
            ("ADF - الفرق الأول - القيمة الاحتمالية", ts["adf_first_diff"]["pvalue"]),
            ("KPSS - المستوى - الإحصائية", ts["kpss_level"]["stat"]),
            ("KPSS - المستوى - القيمة الاحتمالية", ts["kpss_level"]["pvalue"]),
            ("Ljung-Box (تأخر 4) - الإحصائية", ts["ljung_box"][0]["lb_stat"]),
            ("Ljung-Box (تأخر 4) - القيمة الاحتمالية", ts["ljung_box"][0]["lb_pvalue"]),
            ("Ljung-Box (تأخر 8) - الإحصائية", ts["ljung_box"][1]["lb_stat"]),
            ("Ljung-Box (تأخر 8) - القيمة الاحتمالية", ts["ljung_box"][1]["lb_pvalue"]),
            ("Durbin-Watson", ts["durbin_watson"]),
            ("Jarque-Bera - الإحصائية", ts["jarque_bera"]["stat"]),
            ("Jarque-Bera - القيمة الاحتمالية", ts["jarque_bera"]["pvalue"]),
            ("Breusch-Pagan - الإحصائية", ts["breusch_pagan"]["lm_stat"]),
            ("Breusch-Pagan - القيمة الاحتمالية", ts["breusch_pagan"]["lm_pvalue"]),
            ("White Test - الإحصائية", ts["white_test"]["lm_stat"]),
            ("White Test - القيمة الاحتمالية", ts["white_test"]["lm_pvalue"]),
            ("Ramsey RESET - الإحصائية", ts["ramsey_reset"]["stat"]),
            ("Ramsey RESET - القيمة الاحتمالية", ts["ramsey_reset"]["pvalue"]),
            ("Chow Test (يناير 2026) - الإحصائية", ts["chow_test_split_2026_01"]["f_stat"]),
            ("Chow Test (يناير 2026) - القيمة الاحتمالية", ts["chow_test_split_2026_01"]["pvalue"]),
            ("Chow Test (منتصف العينة) - الإحصائية", ts["chow_test_split_midpoint"]["f_stat"]),
            ("Chow Test (منتصف العينة) - القيمة الاحتمالية", ts["chow_test_split_midpoint"]["pvalue"]),
            ("R2 - نموذج الاتجاه والموسمية", ts["ols_trend_seasonal"]["r2"]),
            ("R2 المعدل - نموذج الاتجاه والموسمية", ts["ols_trend_seasonal"]["r2_adj"]),
            ("معامل الاتجاه (ج.م/شهر)", ts["ols_trend_seasonal"]["params"]["trend"]),
            ("القيمة الاحتمالية لمعامل الاتجاه", ts["ols_trend_seasonal"]["pvalues"]["trend"]),
        ]
        write_df(writer, pd.DataFrame(ts_rows, columns=["الاختبار / المعامل", "القيمة"]), "15_الاختبارات الإحصائية")

        # 12. Data quality
        write_df(writer, dict_to_df(dq), "16_جودة البيانات")

        # 13. Full line-item transaction detail (every parsed invoice line)
        tx = pd.read_csv(P / "sales_transactions_enriched.csv")
        write_df(writer, tx, "17_تفاصيل بنود الفواتير الكاملة")

        # reorder: put summary sheet first (already inserted first by write order)
        writer.book.move_sheet("0_الملخص التنفيذي", offset=-len(writer.book.sheetnames))

    print("wrote", OUT, OUT.stat().st_size, "bytes")


if __name__ == "__main__":
    main()
