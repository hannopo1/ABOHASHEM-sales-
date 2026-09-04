#!/usr/bin/env python3
"""Extract the monthly income statements from the costing repo into JSON.

Run manually, not from the pipeline:

    python3 analysis/tools/extract_income_statements.py --repo ../Abohashem

It writes ``data/cost/income_statements.json`` — the same vendoring pattern as
``data/cost/model_rows.json``: a parsed, validated, reviewable extract lives in
this repo, while the binary sources stay in ``hannopo1/Abohashem``.

WHY THE PARSING RULES LOOK PARANOID

The sources are eleven observations spread over one workbook and five PDFs, and
neither format is machine-friendly:

* The workbook's six sheets each start at a different origin (B4, E5, G6, G7,
  I9, G5), so a positional rule reads a different line on every sheet. Worse,
  month 11 carries a stray value in the same row as صافي المبيعات — the naive
  "last number in the row" rule silently returned it. Labels are the only stable
  anchor, and even they need care: the group totals sit in one column per sheet,
  sometimes on the header row and sometimes one row below it.

* The PDFs' text layer is corrupted by Arabic ligature decomposition. The same
  word appears as مصروفات / مرصوفات, صافى / صافن; group headers are truncated
  mid-word (مرصوفات تش, مرصوفات بي); numbers detach from their labels; and in
  one document the group totals are emitted *after* the net-profit line.

The four headline figures — net sales, cost of sales, gross profit, net profit —
are read from the flat text, where they are reliably label-adjacent and
comma-formatted, plus the percentage column where the document prints one.

WHY THE EXPENSE BREAKDOWN IS NOW READ FROM THE PDFs TOO

This extractor used to refuse to parse the expense items out of a PDF, on the
grounds that every general rule tried against these five documents mis-read at
least one, and a plausible wrong number is worse than an absent one. That
refusal was aimed at the FLAT text, and it was right about the flat text: there
the labels glue onto the figures (منظفات ومكافحة7870) and the reading order
scrambles.

Two things changed the answer, and neither is a softening of the standard:

1. The page is read as POSITIONED WORDS (``page.get_text("words")``) instead of
   a flattened stream. Every document turns out to be a clean table once the
   coordinates are kept: the group header sits in its own left-hand column, the
   item label in the middle, the amount in one column, the percentage on the
   right. Nothing has to be inferred from reading order.

2. The documents state three things that must agree, so a misread figure cannot
   pass quietly: each group's own printed total, the sum of the groups against
   gross profit minus net profit, and each item's printed percentage of net
   sales. A month that fails any of them ships with its totals and no items —
   the same fallback as before, now reached by measurement rather than by
   blanket policy.

Total operating expenses is still carried as gross profit minus net profit,
which is exact by construction, and remains the figure everything downstream
adds up to.

Every observation must pass its identities or the run aborts. A statement that
does not reconcile is a statement we do not understand, and guessing which
figure is wrong is exactly how a bad number reaches a board report.
"""
from __future__ import annotations

import argparse
import io
import json
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data" / "cost" / "income_statements.json"

# Absolute tolerance in EGP for every reconciliation below. The statements are
# kept to the piastre, so anything above a pound is a parsing fault, not
# rounding.
TOL = 1.0

# The June 2026 statement is the anchor between the two repositories: these are
# the values analysis/13_join_cost_margin.py already carries as STATEMENT, read
# from this same document. If the extractor disagrees with them, either the
# extraction broke or the source changed — both must stop the run.
JUNE_ANCHOR = {
    "net_sales": 3_741_772.00,
    "cogs": 2_039_933.08,
    "gross_profit": 1_701_838.92,
    "net_profit": 418_841.92,
}

XLSX_SOURCE = "قائمة دخل تفصيلى  من ٧ الى ١٢ سنة ٢٠٢٥.xlsx"

# Sheet name -> period. Taken from the workbook, not inferred from sheet order,
# because the sheet titles carry the month and the order is not guaranteed.
XLSX_SHEETS = {
    "دخل شهر ٧سنة ٢٠٢٥": "2025-07",
    "دخل شهر ٨ سنة ٢٠٢٥": "2025-08",
    "شهر 9سنة ٢٠٢٥": "2025-09",
    "شهر 10سنة ٢٠٢٥": "2025-10",
    "شهر 11سنة ٢٠٢٥": "2025-11",
    "شهر 12سنة ٢٠٢٥": "2025-12",
}

# Each PDF holds one observation. ``months`` is what the period actually spans:
# the first document is a single combined quarterly statement, not three monthly
# ones, and nothing downstream may treat it as monthly without saying so.
PDF_SOURCES = [
    ("قائمة دخل شهر يناير وفبراير ومارس سنة ٢٠٢٦.pdf", "2026-Q1", 3),
    ("دخل شهر 4 سنة 2026.pdf", "2026-04", 1),
    ("قائمة دخل تفصيلى شهر 5 سنة 2026.pdf", "2026-05", 1),
    ("قائمة دخل تفصيلى شهر ٦ سنة ٢٠٢٦.pdf", "2026-06", 1),
    ("قائمة دخل تفصيلى شهر 7 سنه 2026-1.pdf", "2026-07", 1),
]


class ExtractError(SystemExit):
    """Raised for any unreconciled or unreadable statement."""


def die(msg: str) -> None:
    raise ExtractError(f"extract_income_statements: {msg}")


def close(a: float, b: float, tol: float = TOL) -> bool:
    return abs(a - b) <= tol


# --------------------------------------------------------------- git access --

def git_show(repo: Path, ref: str, path: str) -> bytes:
    """Read a file out of the costing repo without checking anything out."""
    p = subprocess.run(["git", "-C", str(repo), "show", f"{ref}:{path}"],
                       capture_output=True)
    if p.returncode != 0:
        die(f"cannot read {path!r} at {ref}: "
            f"{p.stderr.decode('utf-8', 'replace').strip()}")
    return p.stdout


def git_rev(repo: Path, ref: str) -> str:
    p = subprocess.run(["git", "-C", str(repo), "rev-parse", ref],
                       capture_output=True, text=True)
    if p.returncode != 0:
        die(f"cannot resolve {ref!r} in {repo}")
    return p.stdout.strip()


# ------------------------------------------------------------------- xlsx ----

def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _label_at(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v.strip() if isinstance(v, str) else ""


def find_label(ws, *needles):
    """First cell whose text contains every needle. Returns (row, col)."""
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                t = c.value.replace("‏", "")
                if all(n in t for n in needles):
                    return c.row, c.column
    return None


def value_right_of(ws, row, col, max_span=6):
    """Nearest numeric cell to the right of a label, in the same row.

    'Nearest', not 'last': month 11 puts an unrelated figure further right on
    the صافي المبيعات row, and taking the last number reads that instead.
    """
    for j in range(col + 1, col + 1 + max_span):
        v = _num(ws.cell(row=row, column=j).value)
        if v is not None:
            return v
    return None


# A1-style reference, single cell or range, with or without $ anchors.
_REF = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?")


def _cells_in_formula(formula):
    """Every cell a total's formula actually adds up, in order.

    The group totals are live formulas — =SUM(D13:D30), =D49+D50 — so the
    accountant has already declared which rows belong to the group. Reading
    that range beats any bound we could infer: in July the selling total is
    =SUM(D32:D43) while two more expense rows sit at D44 and D45, so a
    "read until the next header" rule silently adds 4,050 the statement never
    counted, and the sheet stops reconciling to its own net profit.
    """
    from openpyxl.utils import range_boundaries
    out = []
    for m in _REF.finditer(formula):
        a = f"{m.group(1)}{m.group(2)}"
        b = f"{m.group(3)}{m.group(4)}" if m.group(3) else a
        c1, r1, c2, r2 = range_boundaries(f"{a}:{b}")
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                out.append((r, c))
    return out


def _label_left_of(ws, row, col, span=3):
    """The nearest non-empty text cell to the left — the item's own name."""
    for c in range(col - 1, max(0, col - 1 - span), -1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip():
            return re.sub(r"\s+", " ", v.strip())
    return ""


def expense_items(ws_val, ws_frm, total_col, header_rows, period):
    """Line items per group, taken from each group total's own formula.

    Returns (items, outside) where `outside` lists expense rows that sit in the
    sheet but in no group's summed range. Those are reported, never folded in:
    the group totals are what tie to net profit, so a row the statement did not
    count is a finding about the source, not a number to add.
    """
    items, counted = [], set()
    for group, (hrow, hcol) in header_rows.items():
        formula = None
        for r in (hrow, hrow + 1):
            v = ws_frm.cell(row=r, column=total_col).value
            if isinstance(v, str) and v.startswith("="):
                formula = v
                break
        if formula is None:
            # A literal total states no membership; ship the group without
            # items rather than guessing which rows it covers.
            return [], []
        for (r, c) in _cells_in_formula(formula):
            counted.add((r, c))
            amount = _num(ws_val.cell(row=r, column=c).value)
            if amount is None:
                continue
            label = _label_left_of(ws_val, r, c)
            # A blank row inside a SUM range that also carries no figure holds
            # nothing to report — the accountant left room to add an item. A
            # blank label with a real figure is the opposite: it is money the
            # statement counted and cannot name, so it is kept and shown
            # unnamed rather than quietly dropped.
            if not label and float(amount) == 0.0:
                continue
            items.append({"group": group, "row": r,
                          "label_raw": label,
                          "amount": float(amount)})

    # Anything in the items column of an expense block that no total counted.
    outside = []
    if items:
        col = {c for _r, c in counted}
        lo, hi = min(r for _g, (r, _c) in header_rows.items()), ws_val.max_row
        for r in range(lo, hi + 1):
            for c in col:
                if (r, c) in counted:
                    continue
                amount = _num(ws_val.cell(row=r, column=c).value)
                label = _label_left_of(ws_val, r, c)
                if amount is None or not label:
                    continue
                if any(k in label for k in ("صاف", "مجمل", "تكلفة", "مصروفات", "مصاريف")):
                    continue
                outside.append({"row": r, "label_raw": label, "amount": float(amount)})
    return items, outside


def parse_sheet(ws, period, source, ws_frm=None):
    """One monthly statement out of one worksheet."""
    anchors = {}
    for key, needles in (("net_sales", ("صافي", "المبيعات")),
                         ("cogs", ("تكلفة", "المبيعات")),
                         ("gross_profit", ("مجمل", "الربح"))):
        hit = find_label(ws, *needles)
        if hit is None:
            die(f"{period}: no {key} label in sheet {ws.title!r}")
        val = value_right_of(ws, *hit)
        if val is None:
            die(f"{period}: {key} label at {hit} has no value to its right")
        anchors[key] = float(val)

    # The expense-group totals all sit in one column per sheet. Establish it
    # from the administrative group, which is well-formed on every sheet, then
    # read the other groups from that same column.
    admin = find_label(ws, "مصروفات", "إدارية")
    if admin is None:
        die(f"{period}: no administrative-expenses header")
    total_col = None
    for j in range(admin[1] + 1, admin[1] + 7):
        if _num(ws.cell(row=admin[0], column=j).value) is not None:
            total_col = j
            break
    if total_col is None:
        die(f"{period}: administrative group has no total")

    groups = {}
    header_rows = {}
    for name, needles in (("admin", ("مصروفات", "إدارية")),
                          ("selling", ("مصروفات", "بيعية")),
                          ("financing", ("مصاريف", "تمويلية"))):
        hit = find_label(ws, *needles)
        if hit is None:
            die(f"{period}: no {name} group header")
        header_rows[name] = hit
        # Month 9's sheet is shifted: its group headers carry a line item and
        # the group total lands one row lower. Accept either row, nothing else.
        val = None
        for r in (hit[0], hit[0] + 1):
            v = _num(ws.cell(row=r, column=total_col).value)
            if v is not None:
                val = float(v)
                break
        if val is None:
            die(f"{period}: {name} group total not in column {total_col}")
        groups[name] = val

    # Net profit sits in the totals column on most sheets, one column left of
    # the label on two of them, and one row above the label on a third. Take
    # the totals column near the label first; fall back to the only number on
    # the label's own row; derive it last, and say so.
    np_hit = find_label(ws, "صاف", "الربح")
    net_profit, np_source = None, "stated"
    if np_hit:
        for r in (np_hit[0], np_hit[0] - 1, np_hit[0] + 1):
            if r < 1:
                continue
            v = _num(ws.cell(row=r, column=total_col).value)
            if v is not None:
                net_profit = float(v)
                break
        if net_profit is None:
            row_vals = [_num(c.value) for c in ws[np_hit[0]]]
            row_vals = [v for v in row_vals if v is not None]
            if len(row_vals) == 1:
                net_profit = float(row_vals[0])
    if net_profit is None:
        net_profit = anchors["gross_profit"] - sum(groups.values())
        np_source = "derived"

    # Line items, when the formula workbook is available. Each group's items
    # must add up to the total that group printed; a group that does not
    # reconcile ships without items rather than with doubtful ones.
    items, outside = ([], [])
    if ws_frm is not None:
        items, outside = expense_items(ws, ws_frm, total_col, header_rows, period)
        by_group = {}
        for it in items:
            by_group[it["group"]] = by_group.get(it["group"], 0.0) + it["amount"]
        for name, total in groups.items():
            got = by_group.get(name, 0.0)
            if items and not close(got, total):
                die(f"{period}: {name} items sum to {got:,.2f} but the sheet's "
                    f"own total says {total:,.2f} — refusing to publish a "
                    f"breakdown that contradicts the statement it came from")

    obs = dict(period=period, months=1, basis="measured",
               source=source, sheet=ws.title,
               net_sales=anchors["net_sales"], cogs=anchors["cogs"],
               gross_profit=anchors["gross_profit"],
               expenses=groups,
               expense_items=items,
               expense_rows_outside_totals=outside,
               total_expenses=sum(groups.values()),
               net_profit=net_profit, net_profit_source=np_source,
               stated_cogs_pct=None, stated_gross_margin_pct=None)
    return obs


def parse_workbook(repo, ref):
    import openpyxl
    raw = git_show(repo, ref, XLSX_SOURCE)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    # A second read keeping formulas: the group totals declare their own item
    # membership, and only the un-evaluated workbook carries that declaration.
    wb_frm = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
    missing = set(XLSX_SHEETS) - set(wb.sheetnames)
    if missing:
        die(f"workbook is missing expected sheets: {sorted(missing)}")
    return [parse_sheet(wb[name], period, XLSX_SOURCE, ws_frm=wb_frm[name])
            for name, period in sorted(XLSX_SHEETS.items(), key=lambda kv: kv[1])]


# -------------------------------------------------------------------- pdf ----

# A money token: comma-grouped, or four or more bare digits. Two- and
# three-digit runs are excluded on purpose — the corrupted text layer glues
# year fragments onto labels (سنة26, ربع1), and those must never be read as
# values.
MONEY = re.compile(r"\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d{4,}(?:\.\d+)?")
PCT = re.compile(r"(\d+(?:\.\d+)?)\s*%")


def norm(text: str) -> str:
    """Collapse the PDF's line breaks and ligature debris into one stream."""
    text = text.replace("‏", " ").replace("‎", " ")
    return re.sub(r"\s+", " ", text)


def after(stream: str, pattern: str, period: str, what: str):
    """The first money token after a label, plus the percentage if one trails.

    Labels are matched loosely because the same word decomposes differently
    across documents: صافى / صافن, and spaces appear inside words (الرب ح).
    """
    m = re.search(pattern, stream)
    if not m:
        die(f"{period}: cannot find the {what} label")
    tail = stream[m.end():m.end() + 120]
    money = MONEY.search(tail)
    if not money:
        die(f"{period}: no value after the {what} label")
    value = float(money.group(0).replace(",", ""))
    # A percentage may sit either side of the figure depending on the document.
    pct = None
    around = tail[:money.end() + 12]
    pm = PCT.search(around)
    if pm:
        pct = float(pm.group(1))
    return value, pct


# ------------------------------------------------- pdf expense line items ----

# The four expense categories the 2026 statements use, in the order the
# documents print them. Not a taxonomy of ours: from May 2026 the statements
# name these groups themselves, and the earlier ones print the first three.
PDF_CATEGORY_ORDER = ["operating", "admin", "selling", "financing"]

# A header word is one of these, sitting in the left-hand column. The spelling
# is mangled — مصروفات renders as مرصوفات — so both are accepted.
_HEAD_WORDS = ("مرصوفات", "مصروفات", "مصاريف")

# What a header's own suffix says about which group it opens. June prints three
# bare «مرصوفات» headers with no suffix at all, so this resolves what it can and
# position fills the rest.
_HEAD_HINTS = [("تشغيل", "operating"), ("تش", "operating"),
               ("إدار", "admin"), ("ادار", "admin"), ("إد", "admin"),
               ("بيع", "selling"), ("بي", "selling"),
               ("تمويل", "financing"), ("تمو", "financing"), ("تم", "financing")]

# A word that can only belong to one group, used to check the assignment above
# rather than trust it. A financing group with no loan in it, or a selling group
# with no vehicle and no commission, means the headers were read in the wrong
# order and every item is in the wrong bucket.
_GROUP_MARKERS = {"selling": ("عربية", "عمولة", "المبيعات"),
                  "financing": ("قرض",),
                  "operating": ("تصنيع", "عمالة", "مرتبات"),
                  "admin": ("محاسبة", "مرتبات", "ايجارات", "محا")}

_STOP = ("صافى", "صافن", "صافي")          # net profit closes the last group
_PCT_TOK = re.compile(r"^\d+(?:\.\d+)?%$")
_NUM_TOK = re.compile(r"^-?\d[\d,]*(?:\.\d+)?$")

# Rows sit in horizontal bands. A band is this many points tall: one printed
# row splits into two bands in several documents (July's مرتبات is 3pt above
# its own amount), and the closest two distinct rows ever come is 12pt.
BAND = 7.0

# A group header is set in its own column, far to the left of the item labels.
# The gap is derived per document rather than fixed: the five documents use
# different page layouts (items begin near x=200 in one and near x=420 in
# another), and a fixed threshold read «مصاريف المكتب» and «مصاريف قرض
# المبادرة» as group headers — item rows that merely start with the same word.
HEAD_GAP = 60.0


def _bands(doc):
    """Every row of every page as (page, y, [(x, word), …]) sorted right→left.

    Right to left because the documents are RTL: the label reads from the
    rightmost word, and the amount column is further right still.
    """
    out = []
    for pi, page in enumerate(doc):
        rows = {}
        for x0, y0, _x1, _y1, text, *_rest in page.get_text("words"):
            key = next((k for k in rows if abs(k - y0) <= BAND), y0)
            rows.setdefault(key, []).append((x0, text))
        for y in sorted(rows):
            out.append((pi, y, sorted(rows[y], key=lambda z: -z[0])))
    return out


# A number this close to the label's own words is part of the name, not a
# figure: «عربية 1639» and «كباس تير يد 3782» are vehicles, and dropping their
# digits would merge four different vehicles into one item called «عربية». The
# page furniture these documents also carry — a stray 0 printed far to the left
# of the label — sits 75 points away or more, so the two never overlap.
LABEL_NUM_GAP = 50.0


def _split(row):
    """(label words, figures, percentages) for one band.

    A percentage is neither. A number inside the label's own span is a label
    word; the rest are figures.
    """
    words = [(x, t) for x, t in row if not _PCT_TOK.match(t)]
    pcts = [float(t[:-1]) for _x, t in row if _PCT_TOK.match(t)]
    text = [(x, t) for x, t in words if not _NUM_TOK.match(t)]
    digits = [(x, t) for x, t in words if _NUM_TOK.match(t)]
    if not digits:
        return text, [], pcts
    # The amount is the rightmost figure; every other number is judged by how
    # close it sits to the name.
    amount = max(digits, key=lambda z: z[0])
    rest = [d for d in digits if d is not amount]
    # The label reads right to left, so a numeric suffix sits just LEFT of its
    # words — inside the name's own span, never out beyond its right edge. That
    # right-hand bound is what keeps a group total (printed far to the right of
    # a header that is a single word) from being read as part of the header.
    if text:
        lo = min(x for x, _t in text) - LABEL_NUM_GAP
        hi = max(x for x, _t in text)
    else:
        lo = hi = None
    part = [] if lo is None else [d for d in rest if lo <= d[0] < hi]
    labels = sorted(text + part, key=lambda z: -z[0])
    nums = [(x, float(t.replace(",", "")))
            for x, t in [amount] + [d for d in rest if d not in part]]
    return labels, nums, pcts


def _amount(nums):
    """The figure in the amount column: the rightmost number on the row.

    Not "any number on the row" — June and July print a stray 0 inside the label
    band of the ايجارات row, and April prints one on تصنيع لدى الغير.
    """
    return max(nums, key=lambda z: z[0])[1] if nums else None


def _subsets(rows, most):
    """Every combination of up to `most` rows, smallest first.

    Used only to identify rows a group total does not count, and only when the
    answer is unique — see the reconciliation below.
    """
    import itertools
    out = []
    for n in range(1, most + 1):
        out.extend(itertools.combinations(rows, n))
    return out


def pdf_expense_items(doc, period, net_sales, total_expenses):
    """Line items, grouped, out of one 2026 income-statement PDF.

    Returns (categories, items, outside) or (None, [], []) when the document
    does not reconcile — in which case the caller ships the month with its
    totals and no breakdown, exactly as an unreadable workbook sheet would.

    WHY THIS IS SAFE TO READ AND THE FLAT TEXT WAS NOT

    The refusal this replaces was aimed at the flattened text layer, where the
    Arabic decomposes and labels glue onto figures (منظفات ومكافحة7870). Read as
    positioned words the same page is a clean table, and — the part that
    actually settles it — the document states three things that must agree:
    each group's printed total, the sum of the groups against gross profit minus
    net profit, and each item's own printed percentage of net sales. A misread
    figure fails at least one. That is the difference between a number that is
    probably right and one that cannot be wrong without being caught.
    """
    rows = _bands(doc)

    # Where this document sets its item labels, measured from the document
    # itself: the first (rightmost) word of a row that carries a figure.
    starts = []
    for _pi, _y, row in rows:
        labels, nums, _p = _split(row)
        if labels and nums:
            starts.append(max(x for x, _t in labels))
    if not starts:
        return None, [], []
    starts.sort()
    head_x = starts[len(starts) // 2] - HEAD_GAP

    # Pass 1: locate the header rows and read what their suffix says.
    heads = []
    for i, (_pi, _y, row) in enumerate(rows):
        labels, _nums, _pcts = _split(row)
        if not labels:
            continue
        text = " ".join(t for _x, t in labels)
        if max(x for x, _t in labels) < head_x and \
           any(w in text for w in _HEAD_WORDS):
            cat = next((c for needle, c in _HEAD_HINTS if needle in text), None)
            heads.append({"i": i, "text": text, "cat": cat})
    if not heads:
        return None, [], []

    # Fill the unresolved ones by position: the documents print the groups in a
    # fixed order, so a bare «مرصوفات» is whichever category has not been used.
    taken = {h["cat"] for h in heads if h["cat"]}
    spare = [c for c in PDF_CATEGORY_ORDER if c not in taken]
    for h in heads:
        if h["cat"] is None:
            if not spare:
                return None, [], []
            h["cat"] = spare.pop(0)
    if len({h["cat"] for h in heads}) != len(heads):
        return None, [], []

    # Pass 2: the rows under each header, up to the next header or net profit.
    stop = len(rows)
    for i, (_pi, _y, row) in enumerate(rows):
        if i > heads[0]["i"] and any(w in t for _x, t in row for w in _STOP):
            stop = i
            break
    bounds = [(h, heads[k + 1]["i"] if k + 1 < len(heads) else stop)
              for k, h in enumerate(heads)]

    items, printed = [], {}
    for h, end in bounds:
        labels, nums, _p = _split(rows[h["i"]][2])
        # The header's own total, when the document prints it beside the header.
        # April prints all three on pages of their own instead; those are
        # matched by value further down.
        printed[h["cat"]] = _amount([n for n in nums if n[0] >= head_x])
        carry = []
        for i in range(h["i"] + 1, end):
            _pi, _y, row = rows[i]
            labels, nums, pcts = _split(row)
            amount = _amount(nums)
            if amount is None:
                # A label with no figure of its own continues the row below it:
                # «الفاسير» and «فتيس» and «عىل» are printed on their own band.
                carry = labels + carry
                continue
            label = " ".join(t for _x, t in sorted(labels + carry,
                                                   key=lambda z: -z[0]))
            carry = []
            if not label and amount == 0:
                continue                  # an empty row, not a nameless figure
            items.append({"group": h["cat"], "row": i,
                          "label_raw": re.sub(r"\s+", " ", label).strip(),
                          "amount": float(amount),
                          "stated_pct": pcts[0] if pcts else None})

    # The assignment must agree with what the items actually are.
    for cat, markers in _GROUP_MARKERS.items():
        mine = [it for it in items if it["group"] == cat]
        if mine and not any(m in it["label_raw"] for it in mine
                            for m in markers):
            die(f"{period}: the group read as {cat} contains none of "
                f"{markers} — the headers were matched in the wrong order")

    # April prints its three group totals on pages of their own, after the
    # net-profit line and with no label beside them. They are still the
    # document's own totals, so they are read where it put them and matched by
    # VALUE rather than by position — position would be a guess, value is not.
    orphans = [_amount(nums) for i, (_pi, _y, row) in enumerate(rows)
               if i >= stop
               for labels, nums, _p in [_split(row)] if nums and not labels]

    # Reconcile each group against its own printed total. A shortfall that is
    # exactly the value of some rows means those rows are printed in the sheet
    # and counted by nothing — the same finding the 2025 workbook produces ten
    # times over. They are reported and never folded in. The subset must be the
    # ONLY one that fits, or the read is not understood and the month ships
    # with its totals alone.
    outside = []
    for cat in sorted({it["group"] for it in items}):
        mine = [it for it in items if it["group"] == cat]
        got = sum(it["amount"] for it in mine)
        want = printed.get(cat)
        wants = [want] if want is not None else orphans
        if any(close(got, w) for w in wants):
            continue
        # Only rows that carry a figure can explain a shortfall: a zero row
        # removed changes nothing, so including it would make every answer
        # ambiguous — {20,600} and {20,600 + a zero} both "fit" and the read
        # would be rejected for a difference that is not one.
        pool = [it for it in mine if it["amount"]]
        fits = [drop for drop in _subsets(pool, 3)
                if any(close(got - sum(d["amount"] for d in drop), w)
                       for w in wants)]
        if len(fits) != 1:
            return None, [], []
        for d in fits[0]:
            outside.append({"row": d["row"], "label_raw": d["label_raw"],
                            "amount": d["amount"]})
        items = [it for it in items if it not in fits[0]]

    groups = {}
    for it in items:
        groups[it["group"]] = groups.get(it["group"], 0.0) + it["amount"]
    for cat, total in groups.items():
        if printed.get(cat) is not None and not close(printed[cat], total):
            return None, [], []
    # The check that ties the whole reading to the statement's own bottom line.
    if not close(sum(groups.values()), total_expenses):
        return None, [], []

    # The third check, and the only one the source gives per line: each item's
    # printed share of net sales.
    if net_sales:
        for it in items:
            if it["stated_pct"] is None:
                continue
            got = it["amount"] / net_sales * 100
            if abs(got - it["stated_pct"]) > 0.02:
                die(f"{period}: «{it['label_raw']}» reads {it['amount']:,.2f} "
                    f"= {got:.2f}% of net sales, but the document prints "
                    f"{it['stated_pct']:.2f}%")

    for it in items:
        it.pop("stated_pct", None)
    return groups, items, outside

def parse_pdf(repo, ref, path, period, months):
    import pymupdf
    raw = git_show(repo, ref, path)
    doc = pymupdf.open(stream=raw, filetype="pdf")
    stream = norm(" ".join(page.get_text() for page in doc))

    net_sales, _ = after(stream, r"صافي\s*المبيعات", period, "net sales")
    cogs, cogs_pct = after(stream, r"تكلفة\s*المبيعات", period, "cost of sales")
    gross, gross_pct = after(stream, r"مجمل\s*الرب\s*ح|مجمل\s*الربح",
                             period, "gross profit")
    net_profit, _ = after(stream, r"صاف[ىني]\s*الرب\s*ح|صاف[ىني]\s*الربح",
                          period, "net profit")

    total_expenses = gross - net_profit
    groups, items, outside = pdf_expense_items(doc, period, net_sales,
                                               total_expenses)

    return dict(period=period, months=months, basis="measured",
                source=path, sheet=None,
                net_sales=net_sales, cogs=cogs, gross_profit=gross,
                # None, not {}, when the document did not reconcile: the month
                # then states one total and says so, rather than showing four
                # zeroes for groups nobody read.
                expenses=groups, expense_items=items,
                expense_rows_outside_totals=outside,
                total_expenses=total_expenses,
                net_profit=net_profit, net_profit_source="stated",
                stated_cogs_pct=cogs_pct, stated_gross_margin_pct=gross_pct)


# ---------------------------------------------------------------- validate ----

def validate(obs):
    """Every identity the documents assert about themselves."""
    p = obs["period"]

    if not close(obs["net_sales"] - obs["cogs"], obs["gross_profit"]):
        die(f"{p}: net sales − cost of sales ≠ gross profit "
            f"({obs['net_sales']} − {obs['cogs']} ≠ {obs['gross_profit']})")

    if not close(obs["gross_profit"] - obs["total_expenses"], obs["net_profit"]):
        die(f"{p}: gross profit − expenses ≠ net profit "
            f"({obs['gross_profit']} − {obs['total_expenses']} "
            f"≠ {obs['net_profit']})")

    if obs["expenses"] is not None:
        s = sum(obs["expenses"].values())
        if not close(s, obs["total_expenses"]):
            die(f"{p}: expense groups sum to {s}, total says "
                f"{obs['total_expenses']}")

    # Where the document prints its own percentage column, it is an independent
    # check on the figures above it — the one cross-check the source gives us.
    if obs["net_sales"]:
        if obs["stated_cogs_pct"] is not None:
            got = obs["cogs"] / obs["net_sales"] * 100
            if abs(got - obs["stated_cogs_pct"]) > 0.02:
                die(f"{p}: cost-of-sales ratio {got:.2f}% contradicts the "
                    f"stated {obs['stated_cogs_pct']:.2f}%")
        if obs["stated_gross_margin_pct"] is not None:
            got = obs["gross_profit"] / obs["net_sales"] * 100
            if abs(got - obs["stated_gross_margin_pct"]) > 0.02:
                die(f"{p}: gross margin {got:.2f}% contradicts the stated "
                    f"{obs['stated_gross_margin_pct']:.2f}%")

    if p == "2026-06":
        for k, want in JUNE_ANCHOR.items():
            if not close(obs[k], want):
                die(f"2026-06: {k} is {obs[k]}, but the costing model and "
                    f"analysis/13_join_cost_margin.py carry {want}. The two "
                    f"repositories no longer agree on the same statement.")


# -------------------------------------------------------------------- main ----

def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--repo", required=True, type=Path,
                    help="path to a clone of hannopo1/Abohashem")
    ap.add_argument("--ref", default="origin/main",
                    help="git ref to read the statements from")
    args = ap.parse_args(argv)

    repo = args.repo.expanduser().resolve()
    if not (repo / ".git").exists():
        die(f"{repo} is not a git repository")
    commit = git_rev(repo, args.ref)

    # Read from the resolved commit, never from args.ref: a ref that advances
    # mid-run would mix revisions and still record one commit as the source.
    rows = parse_workbook(repo, commit)
    for path, period, months in PDF_SOURCES:
        rows.append(parse_pdf(repo, commit, path, period, months))
    rows.sort(key=lambda r: r["period"])

    for obs in rows:
        validate(obs)

    covered = sum(r["months"] for r in rows)
    if covered != 13:
        die(f"expected 13 months of coverage, the sources give {covered}")

    payload = {
        "meta": {
            "source_repo": "hannopo1/Abohashem",
            "source_commit": commit,
            "source_ref": args.ref,
            "extracted_on": date.today().isoformat(),
            "extractor": "analysis/tools/extract_income_statements.py",
            "n_observations": len(rows),
            "months_covered": covered,
            "note": "قوائم دخل على مستوى الشركة. لا تحتوي تكلفة لكل صنف.",
        },
        "observations": rows,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2,
                              sort_keys=True) + "\n", encoding="utf-8")

    print(f"wrote {OUT.relative_to(ROOT)} — {len(rows)} observations, "
          f"{covered} months, all identities reconciled")
    for r in rows:
        gm = r["gross_profit"] / r["net_sales"] * 100 if r["net_sales"] else 0
        print(f"  {r['period']}  {r['months']}mo  "
              f"net {r['net_sales']:>13,.2f}  cogs {r['cogs']:>13,.2f}  "
              f"GM {gm:5.2f}%")
    return 0


if __name__ == "__main__":
    sys.exit(main())
