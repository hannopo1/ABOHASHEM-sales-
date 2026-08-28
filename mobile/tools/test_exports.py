#!/usr/bin/env python3
"""Verify the export feature really produces valid files.

    python3 mobile/tools/test_exports.py [path-to-html]

The XLSX writer in dash-export.js is hand-rolled — a ZIP container plus OOXML
parts, with no library — so it has to be proved rather than assumed. This drives
the real page in Chromium, pulls each generated blob back as base64, and opens
it with the same kind of reader a user would: openpyxl for the workbook, a CSV
parser for the CSV, an XML parser for the SVG.
"""
from __future__ import annotations

import base64
import csv
import io
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

ROOT = Path(__file__).resolve().parents[2]
DEFAULT = ROOT / "mobile" / "dist" / "Abu_Hashem_Mobile_standalone.html"
CHROMIUM = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"

fails: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    """Print the result of a validation check and record failures.
    
    Parameters:
    	name (str): Name of the check.
    	ok (bool): Whether the check passed.
    	detail (str): Optional detail to include in the output.
    """
    print(f"  {'PASS' if ok else 'FAIL'}  {name}{('  — ' + detail) if detail else ''}")
    if not ok:
        fails.append(name)


def run(path: Path) -> int:
    """
    Validate CSV, XLSX, and chart exports from the specified mobile HTML application.
    
    Parameters:
    	path (Path): Path to the HTML file to load.
    
    Returns:
    	int: 0 if all export checks pass, otherwise 1.
    """
    with sync_playwright() as pw:
        b = pw.chromium.launch(executable_path=CHROMIUM)
        page = b.new_page(viewport={"width": 390, "height": 844})
        errs: list[str] = []
        page.on("pageerror", lambda e: errs.append(str(e)))
        page.goto(path.as_uri())
        page.wait_for_selector("#root > div", timeout=30_000)
        page.wait_for_timeout(1500)
        # Switch to the 18-month dataset: it carries the widest catalogue.
        page.locator("text='18 شهرًا'").first.click()
        page.wait_for_timeout(1200)

        cat = page.evaluate(
            "() => { const a = window.__app; const t = a.exportTables();"
            " return t.map(x => ({id:x.id, label:x.label, rows:x.rows.length,"
            " cols:x.columns.length})); }")
        check("export catalogue is populated", len(cat) > 0, f"{len(cat)} tables")
        for t in cat:
            print(f"        {t['label']}  {t['rows']} rows x {t['cols']} cols")

        # ---- CSV ----------------------------------------------------------
        csv_b64 = page.evaluate(
            "() => { const t = window.__app.exportTables()[0];"
            " const s = '\\ufeff' + X.toCSV(t);"
            " const u = new TextEncoder().encode(s); let b = '';"
            " for (let i = 0; i < u.length; i++) b += String.fromCharCode(u[i]);"
            " return btoa(b); }")
        raw = base64.b64decode(csv_b64)
        check("CSV starts with a UTF-8 BOM", raw[:3] == b"\xef\xbb\xbf",
              "Excel needs it to read Arabic")
        rows = list(csv.reader(io.StringIO(raw.decode("utf-8-sig"))))
        check("CSV parses with a header and data", len(rows) > 1 and len(rows[0]) > 1,
              f"{len(rows) - 1} data rows x {len(rows[0])} cols")
        check("CSV header is Arabic, not mojibake",
              any("؀" <= c[0] <= "ۿ" for c in rows[0] if c),
              " | ".join(rows[0][:4]))

        # ---- XLSX ---------------------------------------------------------
        xlsx_b64 = page.evaluate(
            "async () => { const t = window.__app.exportTables();"
            " const blob = X.workbook(t);"
            " const buf = await blob.arrayBuffer(); let s='';"
            " const u = new Uint8Array(buf);"
            " for (let i=0;i<u.length;i++) s += String.fromCharCode(u[i]);"
            " return btoa(s); }")
        data = base64.b64decode(xlsx_b64)
        (ROOT / "mobile" / "dist").mkdir(exist_ok=True)
        out = ROOT / "mobile" / "dist" / "_export_test.xlsx"
        out.write_bytes(data)

        check("XLSX is a valid ZIP", zipfile.is_zipfile(io.BytesIO(data)),
              f"{len(data) / 1024:.0f} KB")
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            bad = z.testzip()
            check("every ZIP entry passes its CRC", bad is None, bad or "")
            names = set(z.namelist())
            need = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml",
                    "xl/_rels/workbook.xml.rels", "xl/styles.xml"}
            check("required OOXML parts present", need <= names,
                  f"{len(names)} entries")

        wb = openpyxl.load_workbook(io.BytesIO(data))
        check("openpyxl opens the workbook", True,
              f"{len(wb.sheetnames)} sheets: " + ", ".join(wb.sheetnames[:3]) + " …")
        ws = wb[wb.sheetnames[0]]
        check("first sheet has rows", ws.max_row > 1 and ws.max_column > 1,
              f"{ws.max_row} x {ws.max_column}")
        hdr = [c.value for c in ws[1]]
        check("workbook header is Arabic",
              any(isinstance(h, str) and any("؀" <= ch <= "ۿ" for ch in h)
                  for h in hdr), " | ".join(str(h) for h in hdr[:4]))
        nums = [c.value for r in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 40))
                for c in r if isinstance(c.value, (int, float))]
        check("numeric cells are numbers, not text", len(nums) > 0,
              f"{len(nums)} numeric cells — Excel can total them")
        check("sheet opens right-to-left", ws.sheet_view.rightToLeft is True)

        # ---- charts -------------------------------------------------------
        page.locator("text='المزيد'").first.click()
        page.wait_for_timeout(400)
        page.locator("text='الربحية'").first.click()
        page.wait_for_timeout(1800)
        n = page.evaluate("() => X.charts().length")
        check("charts are registered while mounted", n > 0, f"{n} live instances")
        png = page.evaluate("() => X.chartPNGDataURL(X.charts()[0], 2)")
        is_png = png.startswith("data:image/png;base64,")
        blob = base64.b64decode(png.split(",", 1)[1]) if is_png else b""
        check("chart PNG is a real PNG, not an SVG data URL",
              is_png and blob[:8] == b"\x89PNG\r\n\x1a\n",
              f"{len(blob) / 1024:.0f} KB" if is_png else png[:40])
        # A PNG of the right size can still be blank if the entry animation has
        # not finished, so assert the series colours are actually on the canvas.
        if is_png:
            from PIL import Image
            im = Image.open(io.BytesIO(blob)).convert("RGB")
            px = set(im.convert("RGB").getdata()) if not hasattr(im, "get_flattened_data") else set(im.get_flattened_data())
            wanted = {(59, 130, 246): "gross-margin line",
                      (16, 185, 129): "operating-margin line"}
            found = [n for c, n in wanted.items() if c in px]
            check("chart PNG contains the plotted series, not just axes",
                  len(found) == len(wanted),
                  f"{len(px)} distinct colours; found: {', '.join(found) or 'none'}")

        svg = page.evaluate("() => X.charts()[0].inst.renderToSVGString()")
        try:
            ET.fromstring(svg)
            check("chart SVG is well-formed XML", True, f"{len(svg) / 1024:.0f} KB")
        except ET.ParseError as e:
            check("chart SVG is well-formed XML", False, str(e))

        check("no JS errors during export", not errs, "; ".join(errs[:3]))
        b.close()

    out.unlink(missing_ok=True)
    print()
    if fails:
        print(f"{len(fails)} check(s) FAILED: {', '.join(fails)}")
        return 1
    print("all export checks passed")
    return 0


if __name__ == "__main__":
    sys.exit(run(Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT))
